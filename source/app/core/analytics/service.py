from __future__ import annotations

import csv
import hashlib
import json
import math
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from app.core.models import utc_now_iso


@dataclass(slots=True)
class AnalyticsConfig:
    time_granularity: str = "week"  # event|day|week|month
    event_types: list[str] | None = None
    date_from: str | None = None  # YYYY-MM-DD
    date_to: str | None = None  # YYYY-MM-DD
    min_samples_for_anomaly: int = 8
    projection_rounds: int = 5
    include_charts: bool = True


class AnalyticsService:
    """Compute and export cross-league statistics for the desktop app."""

    def __init__(self, storage_service: Any) -> None:
        self.storage = storage_service

    def run_analysis(self, output_dir: str | Path, config: AnalyticsConfig | None = None) -> dict[str, Any]:
        cfg = config or AnalyticsConfig()
        rows = self._collect_rows(
            event_types=cfg.event_types,
            date_from=cfg.date_from,
            date_to=cfg.date_to,
        )

        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        charts_path = output_path / "charts"
        charts_path.mkdir(parents=True, exist_ok=True)

        overall = self._overall_metrics(rows)
        by_opponent_deck = self._group_metrics(rows, "opponent_deck")
        by_opponent_archetype = self._group_metrics(rows, "opponent_archetype")
        by_event_type = self._group_metrics(rows, "event_type")
        by_my_deck = self._group_metrics(rows, "my_deck")
        by_my_archetype = self._group_metrics(rows, "my_archetype")
        trends = self._trend_metrics(rows, cfg.time_granularity)
        anomalies = self._detect_anomalies(rows, cfg.min_samples_for_anomaly)
        projections = self._record_projections(overall["winrate_pct"] / 100.0, cfg.projection_rounds)

        payload: dict[str, Any] = {
            "generated_at": utc_now_iso(),
            "filters": {
                "event_types": cfg.event_types or [],
                "date_from": cfg.date_from or "",
                "date_to": cfg.date_to or "",
                "time_granularity": cfg.time_granularity,
                "projection_rounds": cfg.projection_rounds,
            },
            "overall": overall,
            "by_opponent_deck": by_opponent_deck,
            "by_opponent_archetype": by_opponent_archetype,
            "by_event_type": by_event_type,
            "by_my_deck": by_my_deck,
            "by_my_archetype": by_my_archetype,
            "trends": trends,
            "anomalies": anomalies,
            "record_projection": projections,
        }

        charts: list[Path] = []
        if cfg.include_charts:
            charts = self._build_charts(payload, charts_path)

        self._write_json(output_path / "analysis.json", payload)
        self._write_markdown(output_path / "analysis.md", payload)
        self._write_tables_csv(output_path, payload)
        self._write_excel(output_path / "analysis.xlsx", payload, charts)

        return {
            "output_dir": str(output_path),
            "files": {
                "json": str(output_path / "analysis.json"),
                "markdown": str(output_path / "analysis.md"),
                "excel": str(output_path / "analysis.xlsx"),
                "csv_dir": str(output_path),
                "charts_dir": str(charts_path),
            },
            "summary": {
                "matches": overall["matches"],
                "wins": overall["wins"],
                "losses": overall["losses"],
                "winrate_pct": overall["winrate_pct"],
                "ci95_low": overall["ci95_low"],
                "ci95_high": overall["ci95_high"],
            },
        }

    def compare_events(
        self,
        event_a: str,
        event_b: str,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> dict[str, Any]:
        a = event_a.strip()
        b = event_b.strip()
        if not a or not b:
            raise ValueError("Both event names are required.")
        if a.lower() == b.lower():
            raise ValueError("Event A and Event B must be different.")

        rows_a = self._collect_rows(event_types=[a], date_from=date_from, date_to=date_to)
        rows_b = self._collect_rows(event_types=[b], date_from=date_from, date_to=date_to)

        overall_a = self._overall_metrics(rows_a)
        overall_b = self._overall_metrics(rows_b)

        return {
            "event_a": a,
            "event_b": b,
            "date_from": date_from or "",
            "date_to": date_to or "",
            "summary_a": overall_a,
            "summary_b": overall_b,
            "delta": {
                "matches": int(overall_a["matches"]) - int(overall_b["matches"]),
                "winrate_pp": round(float(overall_a["winrate_pct"]) - float(overall_b["winrate_pct"]), 1),
                "mulligan_rate_pp": round(
                    float(overall_a["mulligan_rate_pct"]) - float(overall_b["mulligan_rate_pct"]),
                    1,
                ),
                "mana_screw_rate_pp": round(
                    float(overall_a["mana_screw_rate_pct"]) - float(overall_b["mana_screw_rate_pct"]),
                    1,
                ),
                "mana_flood_rate_pp": round(
                    float(overall_a["mana_flood_rate_pct"]) - float(overall_b["mana_flood_rate_pct"]),
                    1,
                ),
            },
        }

    def _collect_rows(
        self,
        event_types: list[str] | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        wanted_events = {x.strip().lower() for x in (event_types or []) if x.strip()}
        from_date = self._parse_optional_date(date_from)
        to_date = self._parse_optional_date(date_to)

        for league in self.storage.list_leagues():
            league_path = str(league.get("path", ""))
            if not league_path:
                continue

            event_type = str(league.get("event_type", "")).strip()
            if wanted_events and event_type.lower() not in wanted_events:
                continue

            snapshot = self.storage.get_league_snapshot(league_path)
            meta = snapshot.get("meta", {})
            matches = snapshot.get("matches", [])
            league_date = self._parse_date(str(meta.get("date", "")))

            if from_date and league_date < from_date:
                continue
            if to_date and league_date > to_date:
                continue

            my_deck = str(meta.get("deck_name", "")).strip() or "Unknown"
            my_archetype = str(meta.get("deck_archetype", "")).strip() or "Unknown"

            for match in matches:
                match_result = str(match.get("match_result", "loss")).strip().lower()
                wins = 1 if match_result == "win" else 0
                losses = 0 if match_result == "win" else 1
                games = list(match.get("games", []))

                mulligan_games = 0
                mana_screw_games = 0
                mana_flood_games = 0
                for game in games:
                    if int(game.get("mulligan_count", 0) or 0) > 0:
                        mulligan_games += 1
                    draw_type = str(game.get("draw_type", "")).strip()
                    if draw_type == "Mana Screw":
                        mana_screw_games += 1
                    elif draw_type == "Mana Flood":
                        mana_flood_games += 1

                rows.append(
                    {
                        "league_id": str(meta.get("league_id", "")),
                        "date": league_date,
                        "event_type": event_type or "Unknown",
                        "my_deck": my_deck,
                        "my_archetype": my_archetype,
                        "opponent_deck": str(match.get("opponent_deck", "")).strip() or "Unknown",
                        "opponent_archetype": str(match.get("opponent_archetype", "")).strip() or "Unknown",
                        "wins": wins,
                        "losses": losses,
                        "games": len(games),
                        "mulligan_games": mulligan_games,
                        "mana_screw_games": mana_screw_games,
                        "mana_flood_games": mana_flood_games,
                    }
                )

        return rows

    def _overall_metrics(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        matches = len(rows)
        wins = sum(int(r["wins"]) for r in rows)
        losses = sum(int(r["losses"]) for r in rows)
        games = sum(int(r["games"]) for r in rows)
        mulligan_games = sum(int(r["mulligan_games"]) for r in rows)
        mana_screw_games = sum(int(r["mana_screw_games"]) for r in rows)
        mana_flood_games = sum(int(r["mana_flood_games"]) for r in rows)
        winrate = self._pct(wins, matches)
        ci_low, ci_high = self._wilson_ci(wins, matches)

        return {
            "matches": matches,
            "wins": wins,
            "losses": losses,
            "winrate_pct": winrate,
            "ci95_low": round(ci_low * 100.0, 1),
            "ci95_high": round(ci_high * 100.0, 1),
            "mulligan_rate_pct": self._pct(mulligan_games, games),
            "mana_screw_rate_pct": self._pct(mana_screw_games, games),
            "mana_flood_rate_pct": self._pct(mana_flood_games, games),
            "games": games,
        }

    def _group_metrics(self, rows: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
        groups: dict[str, dict[str, int]] = {}
        for row in rows:
            label = str(row.get(key, "Unknown") or "Unknown")
            bucket = groups.setdefault(
                label,
                {
                    "matches": 0,
                    "wins": 0,
                    "losses": 0,
                    "games": 0,
                    "mulligan_games": 0,
                    "mana_screw_games": 0,
                    "mana_flood_games": 0,
                },
            )
            bucket["matches"] += 1
            bucket["wins"] += int(row["wins"])
            bucket["losses"] += int(row["losses"])
            bucket["games"] += int(row["games"])
            bucket["mulligan_games"] += int(row["mulligan_games"])
            bucket["mana_screw_games"] += int(row["mana_screw_games"])
            bucket["mana_flood_games"] += int(row["mana_flood_games"])

        result: list[dict[str, Any]] = []
        for label, values in sorted(groups.items(), key=lambda x: x[0].lower()):
            ci_low, ci_high = self._wilson_ci(values["wins"], values["matches"])
            result.append(
                {
                    "name": label,
                    "matches": values["matches"],
                    "wins": values["wins"],
                    "losses": values["losses"],
                    "winrate_pct": self._pct(values["wins"], values["matches"]),
                    "ci95_low": round(ci_low * 100.0, 1),
                    "ci95_high": round(ci_high * 100.0, 1),
                    "mulligan_rate_pct": self._pct(values["mulligan_games"], values["games"]),
                    "mana_screw_rate_pct": self._pct(values["mana_screw_games"], values["games"]),
                    "mana_flood_rate_pct": self._pct(values["mana_flood_games"], values["games"]),
                }
            )

        return sorted(result, key=lambda x: (-int(x["matches"]), str(x["name"]).lower()))

    def _trend_metrics(self, rows: list[dict[str, Any]], granularity: str) -> list[dict[str, Any]]:
        if granularity not in {"event", "day", "week", "month"}:
            granularity = "week"

        buckets: dict[str, dict[str, int]] = {}
        for row in rows:
            label = self._trend_bucket_label(row, granularity)
            bucket = buckets.setdefault(label, {"matches": 0, "wins": 0, "losses": 0})
            bucket["matches"] += 1
            bucket["wins"] += int(row["wins"])
            bucket["losses"] += int(row["losses"])

        result: list[dict[str, Any]] = []
        for label, values in sorted(buckets.items()):
            ci_low, ci_high = self._wilson_ci(values["wins"], values["matches"])
            result.append(
                {
                    "bucket": label,
                    "matches": values["matches"],
                    "wins": values["wins"],
                    "losses": values["losses"],
                    "winrate_pct": self._pct(values["wins"], values["matches"]),
                    "ci95_low": round(ci_low * 100.0, 1),
                    "ci95_high": round(ci_high * 100.0, 1),
                }
            )
        return result

    def _detect_anomalies(self, rows: list[dict[str, Any]], min_samples: int) -> list[dict[str, Any]]:
        by_deck = self._group_metrics(rows, "opponent_deck")
        baseline = self._overall_metrics(rows)
        baseline_wr = float(baseline["winrate_pct"])

        anomalies: list[dict[str, Any]] = []
        for deck in by_deck:
            matches = int(deck["matches"])
            if matches < min_samples:
                continue
            wr = float(deck["winrate_pct"])
            delta = round(wr - baseline_wr, 1)
            if abs(delta) < 20.0:
                continue
            anomalies.append(
                {
                    "scope": "opponent_deck",
                    "name": deck["name"],
                    "matches": matches,
                    "winrate_pct": wr,
                    "delta_vs_overall_pct": delta,
                    "severity": "high" if abs(delta) >= 30.0 else "medium",
                }
            )

        return sorted(anomalies, key=lambda x: (-abs(float(x["delta_vs_overall_pct"])), str(x["name"])))

    def _record_projections(self, winrate: float, rounds: int) -> list[dict[str, Any]]:
        rounds = max(1, int(rounds))
        winrate = max(0.0, min(1.0, float(winrate)))

        rows: list[dict[str, Any]] = []
        for wins in range(rounds + 1):
            losses = rounds - wins
            probability = math.comb(rounds, wins) * (winrate**wins) * ((1.0 - winrate) ** losses)
            rows.append(
                {
                    "record": f"{wins}-{losses}",
                    "probability_pct": round(probability * 100.0, 2),
                }
            )
        return rows

    def _write_json(self, path: Path, payload: dict[str, Any]) -> None:
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    def _write_markdown(self, path: Path, payload: dict[str, Any]) -> None:
        overall = payload["overall"]
        lines = [
            "# Statistics Analysis",
            "",
            "## Overall",
            "",
            f"- Matches: {overall['matches']}",
            f"- Record: {overall['wins']}-{overall['losses']}",
            f"- Winrate: {overall['winrate_pct']:.1f}%",
            f"- 95% CI: {overall['ci95_low']:.1f}% - {overall['ci95_high']:.1f}%",
            f"- Mulligan Rate: {overall['mulligan_rate_pct']:.1f}%",
            f"- Mana Screw Rate: {overall['mana_screw_rate_pct']:.1f}%",
            f"- Mana Flood Rate: {overall['mana_flood_rate_pct']:.1f}%",
            "",
            "## Trends",
            "",
        ]

        trends = payload.get("trends", [])
        if not trends:
            lines.append("- No trend data")
        else:
            for row in trends:
                lines.append(
                    f"- {row['bucket']}: {row['wins']}-{row['losses']} ({row['winrate_pct']:.1f}% WR, CI {row['ci95_low']:.1f}-{row['ci95_high']:.1f})"
                )

        anomalies = payload.get("anomalies", [])
        lines.extend(["", "## Anomalies", ""])
        if not anomalies:
            lines.append("- No anomalies detected")
        else:
            for row in anomalies:
                lines.append(
                    f"- {row['name']}: {row['winrate_pct']:.1f}% (delta {row['delta_vs_overall_pct']:+.1f} pp, n={row['matches']})"
                )

        lines.extend(["", "## Record Projection", ""])
        for row in payload.get("record_projection", []):
            lines.append(f"- {row['record']}: {row['probability_pct']:.2f}%")

        path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def _write_tables_csv(self, output_dir: Path, payload: dict[str, Any]) -> None:
        self._write_csv(output_dir / "overall.csv", [payload["overall"]])
        self._write_csv(output_dir / "by_opponent_deck.csv", payload.get("by_opponent_deck", []))
        self._write_csv(output_dir / "by_opponent_archetype.csv", payload.get("by_opponent_archetype", []))
        self._write_csv(output_dir / "by_event_type.csv", payload.get("by_event_type", []))
        self._write_csv(output_dir / "by_my_deck.csv", payload.get("by_my_deck", []))
        self._write_csv(output_dir / "by_my_archetype.csv", payload.get("by_my_archetype", []))
        self._write_csv(output_dir / "trends.csv", payload.get("trends", []))
        self._write_csv(output_dir / "anomalies.csv", payload.get("anomalies", []))
        self._write_csv(output_dir / "record_projection.csv", payload.get("record_projection", []))

    def _write_csv(self, path: Path, rows: list[dict[str, Any]]) -> None:
        if not rows:
            path.write_text("\n", encoding="utf-8")
            return
        fieldnames = list(rows[0].keys())
        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    def _write_excel(self, path: Path, payload: dict[str, Any], charts: list[Path]) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as XLImage
        except Exception:
            # If openpyxl is not available, create a marker text file instead of failing the flow.
            path.with_suffix(".txt").write_text(
                "Excel export skipped: missing openpyxl dependency.\n",
                encoding="utf-8",
            )
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "overall"
        self._fill_sheet(ws, [payload["overall"]])

        for name in [
            "by_opponent_deck",
            "by_opponent_archetype",
            "by_event_type",
            "by_my_deck",
            "by_my_archetype",
            "trends",
            "anomalies",
            "record_projection",
        ]:
            ws_n = wb.create_sheet(title=name[:31])
            self._fill_sheet(ws_n, payload.get(name, []))

        if charts:
            ws_charts = wb.create_sheet(title="charts")
            current_row = 1
            for chart in charts:
                try:
                    ws_charts.add_image(XLImage(str(chart)), f"A{current_row}")
                    current_row += 20
                except Exception:
                    continue

        wb.save(path)

    def _fill_sheet(self, ws: Any, rows: list[dict[str, Any]]) -> None:
        if not rows:
            ws["A1"] = "No data"
            return

        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header))

    def _build_charts(self, payload: dict[str, Any], charts_dir: Path) -> list[Path]:
        try:
            import matplotlib.pyplot as plt
            from matplotlib.patches import Patch
        except Exception:
            return []

        plt.style.use("seaborn-v0_8-whitegrid")
        chart_paths: list[Path] = []

        trends = payload.get("trends", [])
        if trends:
            x = [str(r["bucket"]) for r in trends]
            y = [float(r["winrate_pct"]) for r in trends]
            ci_low = [float(r.get("ci95_low", 0.0)) for r in trends]
            ci_high = [float(r.get("ci95_high", 0.0)) for r in trends]
            x_idx = list(range(len(x)))

            fig, ax = plt.subplots(figsize=(11, 5))
            ax.plot(x_idx, y, marker="o", linewidth=2.4, color="#1f77b4", label="Winrate")
            ax.fill_between(x_idx, ci_low, ci_high, color="#1f77b4", alpha=0.18, label="95% CI")
            ax.set_title("Winrate Trend")
            ax.set_ylabel("Winrate %")
            ax.set_xlabel("Bucket")
            ax.set_xticks(x_idx)
            ax.set_xticklabels(x, rotation=30, ha="right")
            ax.set_ylim(0, 100)
            ax.legend(bbox_to_anchor=(1.02, 1), loc="upper left", frameon=True)
            ax.grid(alpha=0.25)
            p = charts_dir / "winrate_trend.png"
            fig.tight_layout()
            fig.savefig(p, dpi=150, bbox_inches="tight")
            plt.close(fig)
            chart_paths.append(p)

            y2 = [int(r["matches"]) for r in trends]
            colors = [self._color_for_label(label) for label in x]
            fig2, ax2 = plt.subplots(figsize=(11, 5))
            ax2.bar(x_idx, y2, color=colors)
            ax2.set_title("Matches by Bucket")
            ax2.set_ylabel("Matches")
            ax2.set_xlabel("Bucket")
            ax2.set_xticks(x_idx)
            ax2.set_xticklabels(x, rotation=30, ha="right")
            ax2.grid(axis="y", alpha=0.25)
            legend_labels = x[:10]
            handles = [Patch(facecolor=self._color_for_label(lbl), label=lbl) for lbl in legend_labels]
            if len(x) > 10:
                handles.append(Patch(facecolor="#cccccc", label="..."))
            ax2.legend(handles=handles, title="Buckets",
                       bbox_to_anchor=(1.02, 1), loc="upper left", frameon=True)
            p2 = charts_dir / "matches_by_bucket.png"
            fig2.tight_layout()
            fig2.savefig(p2, dpi=170, bbox_inches="tight")
            plt.close(fig2)
            chart_paths.append(p2)

        top_decks = payload.get("by_opponent_deck", [])[:10]
        if top_decks:
            labels = [str(r["name"]) for r in top_decks]
            wr = [float(r["winrate_pct"]) for r in top_decks]
            counts = [int(r["matches"]) for r in top_decks]
            deck_colors = [self._color_for_label(lbl) for lbl in labels]
            n = len(labels)
            bar_width = max(0.18, min(0.75, 9.0 / max(n, 1)))
            fig_width = min(16, max(12, n * 0.6 + 8))
            tick_font = 9 if n <= 10 else 8
            fig3, ax3 = plt.subplots(figsize=(fig_width, 6))
            x_idx = list(range(n))
            bars3 = ax3.bar(x_idx, wr, width=bar_width, color=deck_colors)
            for bar, count in zip(bars3, counts):
                ax3.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 1.5,
                    f"n={count}",
                    ha="center", va="bottom", fontsize=8, color="#333333",
                )
            ax3.set_title("Top Opponent Decks by Sample")
            ax3.set_xlabel("Deck")
            ax3.set_ylabel("Winrate %")
            ax3.set_ylim(0, 110)
            ax3.set_xticks(x_idx)
            ax3.set_xticklabels(labels, rotation=35, ha="right", fontsize=tick_font)
            ax3.grid(axis="y", alpha=0.25)
            deck_handles = [
                Patch(facecolor=color, label=f"{label} (n={count})")
                for label, color, count in zip(labels, deck_colors, counts)
            ]
            ax3.legend(handles=deck_handles, title="Decks",
                       bbox_to_anchor=(1.02, 1), loc="upper left", frameon=True)
            p3 = charts_dir / "top_opponent_decks_wr.png"
            fig3.tight_layout()
            fig3.savefig(p3, dpi=170, bbox_inches="tight")
            plt.close(fig3)
            chart_paths.append(p3)

        deck_share_rows = payload.get("by_opponent_deck", [])
        if deck_share_rows:
            labels = [str(r["name"]) for r in deck_share_rows]
            counts = [int(r["matches"]) for r in deck_share_rows]
            total_matches = sum(counts)
            if total_matches > 0:
                shares = [(count / total_matches) * 100.0 for count in counts]
                colors = [self._color_for_label(label) for label in labels]

                fig5, ax5 = plt.subplots(figsize=(12, 8))
                wedges, _ = ax5.pie(
                    counts,
                    colors=colors,
                    startangle=90,
                    wedgeprops={"linewidth": 1.0, "edgecolor": "white"},
                )

                points: list[dict[str, Any]] = []
                for wedge, label, share in zip(wedges, labels, shares):
                    angle = (wedge.theta2 + wedge.theta1) / 2.0
                    angle_rad = math.radians(angle)
                    x = math.cos(angle_rad)
                    y = math.sin(angle_rad)
                    side = 1 if x >= 0 else -1
                    points.append(
                        {
                            "label": label,
                            "share": share,
                            "angle": angle,
                            "x": x,
                            "y": y,
                            "side": side,
                            "target_y": 1.15 * y,
                        }
                    )

                def _spread_targets(items: list[dict[str, Any]]) -> None:
                    if not items:
                        return
                    items.sort(key=lambda it: float(it["target_y"]))
                    min_gap = 0.09
                    y_min, y_max = -1.25, 1.25
                    prev_y = y_min - min_gap
                    for item in items:
                        target = float(item["target_y"])
                        target = max(target, prev_y + min_gap)
                        item["target_y"] = target
                        prev_y = target

                    overflow = float(items[-1]["target_y"]) - y_max
                    if overflow > 0:
                        for item in items:
                            item["target_y"] = float(item["target_y"]) - overflow

                    underflow = y_min - float(items[0]["target_y"])
                    if underflow > 0:
                        for item in items:
                            item["target_y"] = float(item["target_y"]) + underflow

                right_side = [p for p in points if int(p["side"]) > 0]
                left_side = [p for p in points if int(p["side"]) < 0]
                _spread_targets(right_side)
                _spread_targets(left_side)

                for item in points:
                    side = int(item["side"])
                    align = "left" if side > 0 else "right"

                    # External labels with connector lines keep the pie readable.
                    ax5.annotate(
                        f"{item['label']} {float(item['share']):.1f}%",
                        xy=(0.92 * float(item["x"]), 0.92 * float(item["y"])),
                        xytext=(1.35 * side, float(item["target_y"])),
                        ha=align,
                        va="center",
                        fontsize=8,
                        arrowprops={
                            "arrowstyle": "-",
                            "color": "#666666",
                            "connectionstyle": f"angle,angleA=0,angleB={float(item['angle'])}",
                            "shrinkA": 0,
                            "shrinkB": 0,
                        },
                    )

                ax5.set_title("Opponent Deck Metagame Share (%)", pad=24)
                ax5.axis("equal")
                p5 = charts_dir / "opponent_deck_share_pie.png"
                fig5.tight_layout()
                fig5.savefig(p5, dpi=170, bbox_inches="tight")
                plt.close(fig5)
                chart_paths.append(p5)

        top_archetypes = payload.get("by_opponent_archetype", [])[:10]
        if top_archetypes:
            labels = [str(r["name"]) for r in top_archetypes]
            wr = [float(r["winrate_pct"]) for r in top_archetypes]
            counts = [int(r["matches"]) for r in top_archetypes]
            archetype_colors = [self._color_for_label(lbl) for lbl in labels]
            n = len(labels)
            bar_width = max(0.18, min(0.75, 9.0 / max(n, 1)))
            fig_width = min(16, max(12, n * 0.6 + 8))
            tick_font = 9 if n <= 10 else 8
            fig4, ax4 = plt.subplots(figsize=(fig_width, 6))
            x_idx = list(range(n))
            bars4 = ax4.bar(x_idx, wr, width=bar_width, color=archetype_colors)
            for bar, count in zip(bars4, counts):
                ax4.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 1.5,
                    f"n={count}",
                    ha="center", va="bottom", fontsize=8, color="#333333",
                )
            ax4.set_title("Top Opponent Archetypes by Sample")
            ax4.set_xlabel("Archetype")
            ax4.set_ylabel("Winrate %")
            ax4.set_ylim(0, 110)
            ax4.set_xticks(x_idx)
            ax4.set_xticklabels(labels, rotation=35, ha="right", fontsize=tick_font)
            ax4.grid(axis="y", alpha=0.25)
            arch_handles = [
                Patch(facecolor=color, label=f"{label} (n={count})")
                for label, color, count in zip(labels, archetype_colors, counts)
            ]
            ax4.legend(handles=arch_handles, title="Archetypes",
                       bbox_to_anchor=(1.02, 1), loc="upper left", frameon=True)
            p4 = charts_dir / "top_opponent_archetypes_wr.png"
            fig4.tight_layout()
            fig4.savefig(p4, dpi=170, bbox_inches="tight")
            plt.close(fig4)
            chart_paths.append(p4)

        return chart_paths

    def _color_for_label(self, label: str) -> tuple[float, float, float]:
        """Stable color assignment based on label text."""
        digest = hashlib.md5(label.encode("utf-8")).hexdigest()
        r = int(digest[0:2], 16) / 255.0
        g = int(digest[2:4], 16) / 255.0
        b = int(digest[4:6], 16) / 255.0

        # Keep colors vivid enough for charts and avoid too-dark bars.
        min_v = 0.25
        r = max(r, min_v)
        g = max(g, min_v)
        b = max(b, min_v)
        return (r, g, b)

    def _trend_bucket_label(self, row: dict[str, Any], granularity: str) -> str:
        if granularity == "event":
            return str(row.get("event_type", "Unknown"))

        dt = row.get("date")
        if not isinstance(dt, date):
            return "unknown"

        if granularity == "day":
            return dt.isoformat()
        if granularity == "month":
            return f"{dt.year:04d}-{dt.month:02d}"

        iso_year, iso_week, _ = dt.isocalendar()
        return f"{iso_year:04d}-W{iso_week:02d}"

    def _parse_date(self, raw: str) -> date:
        value = raw.strip()
        if not value:
            return datetime.now(UTC).date()

        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d %m %Y", "%d.%m.%Y"]:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

        try:
            return datetime.fromisoformat(value).date()
        except Exception:
            return datetime.now(UTC).date()

    def _parse_optional_date(self, raw: str | None) -> date | None:
        value = str(raw or "").strip()
        if not value:
            return None
        return self._parse_date(value)

    def _pct(self, part: int, whole: int) -> float:
        if whole <= 0:
            return 0.0
        return round((float(part) / float(whole)) * 100.0, 1)

    def _wilson_ci(self, wins: int, n: int, z: float = 1.96) -> tuple[float, float]:
        if n <= 0:
            return (0.0, 0.0)
        p_hat = wins / n
        denominator = 1.0 + (z * z) / n
        center = p_hat + (z * z) / (2.0 * n)
        margin = z * math.sqrt((p_hat * (1.0 - p_hat) / n) + ((z * z) / (4.0 * n * n)))
        low = (center - margin) / denominator
        high = (center + margin) / denominator
        return (max(0.0, low), min(1.0, high))
